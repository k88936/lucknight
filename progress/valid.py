import sys
import os
import re
import openpyxl

def validate_filename(filename):
    # 检查文件名格式：大作业自评表_学号_姓名.xlsx，其中学号是数字
    base = os.path.basename(filename)
    pattern = r'^大作业自评表_\d+_.+\.xlsx$'
    if re.match(pattern, base):
        return True, ""
    else:
        return False, "文件名格式不正确，应为：大作业自评表_学号_姓名.xlsx，且学号应为纯数字。"

def validate_content(filename):
    try:
        wb = openpyxl.load_workbook(filename)
        sheet = wb.active

        # 查找“是否实现”列
        header_row = 1
        target_col_index = None
        for col in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(row=header_row, column=col).value
            if cell_value and str(cell_value).strip() == '是否实现':
                target_col_index = col
                break

        if target_col_index is None:
            return False, "缺少“是否实现”列。"

        # 检查从第2行开始的每一行值是否为0或1
        for row in range(2, sheet.max_row + 1):
            cell = sheet.cell(row=row, column=target_col_index)
            value = cell.value
            if value not in (0, 1):
                return False, f"第 {row} 行“是否实现”列的值不是 0 或 1（当前值：{value}）。"

        return True, ""
    except Exception as e:
        return False, f"读取 Excel 文件失败：{str(e)}"

if __name__ == '__main__':
    if len(sys.argv) != 2:
        print("用法错误：请提供一个文件路径作为参数，例如：python valid.py 大作业自评表_2021312589_明鑫.xlsx")
        sys.exit(1)

    filepath = sys.argv[1]
    valid_name, name_msg = validate_filename(filepath)
    valid_content, content_msg = validate_content(filepath)

    if valid_name and valid_content:
        print("符合提交要求！")
    else:
        print("不符合提交要求，原因如下：")
        if not valid_name:
            print(" -", name_msg)
        if not valid_content:
            print(" -", content_msg)
